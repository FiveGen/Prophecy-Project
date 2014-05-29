#INITIAL IMPORTS
from datetime import datetime
import time, random
import facebook
import json

#IMPORT EXCEL READ / WRITE / UTILIS LIBRARIES
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

#DEFINE TIME
pingtime = datetime.now().strftime('%Y-%m-%d %H:%M')
in_time = datetime.now().replace(microsecond=0)

#TITLE
print ' _    _  _____  ______ ___  _____ '
print '| |  | ||  ___||___  //   ||  _  |'
print '| |  | ||___ \    / // /| | \ V / '
print '| |/\| |    \ \  / // /_| | / _ \ '
print '\  /\  //\__/ /_/ / \___  || |_| |'
print ' \/  \/ \____/ \_/      |_/\_____/\n'

print '----------------------------> INFO'
print '* ' + pingtime
print '* \output - facebook.xls' + '\n'
print '----------------------------> LOG'

#OPEN OUTPUT.XLS
wb = load_workbook('output - facebook.xlsx')
ws = wb.get_sheet_by_name("sheet 1")

def pp(o): 
    print json.dumps(o, indent=1)

# Create a connection to the Graph API with your access token
g = facebook.GraphAPI('CAADbHGZCkCBEBADZCJAtRe9BKbg7Irc0JJQmtO7h9nPUI3Ejs7naE2lNkdwJCIjHVZCVSZCsmpasnkkmoeGpoZA0zqRMk374D2JVDoSNxMH38L6e3VYt7OQrnuIrOwIPm12F95YgluskGKMgayM9uIMRRe5LZA3xtjgIvS5PFd7tOWIAcoOkXT')

#PROFILE IDS
LOL = '82061850555'
LOLN = 'League of Legends'

HON = '63037549101'
HONN = 'Heroes of Newerth'

DO2 = '106876872711112'
DO2N = 'Defense of the Ancients 2'

GW2 = '114036714208'
GW2N = 'Guild Wars 2'

XIV = '116214575870'
XIVN = 'Final Fantasy XIV'

WOW = '138011799033'
WOWN = 'World of Warcraft'

ESO = '401899346486771'
ESON = 'Elder Scrolls Online'

EQN = '142045689307777'
EQNN = 'EverQuest Next'

WIS = '138242926249490'
WISN = 'WildStar'

#LIST FORM / POSITION
LIST = [LOL, HON, DO2, GW2, XIV, WOW, ESO, EQN, WIS]
LISTN = [LOLN, HONN, DO2N, GW2N, XIVN, WOWN, ESON, EQNN, WISN]
N = 0

#EXCEL PARAMETERS
R = 1
C = 1
cell = ws.cell(row = R, column = C)

#REPEAT FOR LOOP
repeat = 'yes'
while N <= 8:

    #MOVE THE ROW
    try:
        while repeat == 'yes':
            cell = ws.cell(row = R, column = 1)
            if cell.data_type != 's':
                R += 1
            else:
                repeat = 'no'

    except IndexError:
        pass

    def int_format(y): return "{:}".format(y)
    x = int_format(g.get_object(LIST[N])['likes'])
    i = int(x)
    
    print '[' + str(N) + '] ' + LISTN[N]
    print ' + PROFILE ID: ' + LIST[N]
    print ' + LIKES RETREIVED: ' + str(x) + '\n'
    N+=1

    #WRITE TO THE SPREADSHEET, MOVE THE COLUMN
    target_cell = ws.cell(row=R, column=C)
    target_cell.value = i
    C += 1

date_cell = ws.cell(row=R, column=0)
date_cell.value = pingtime

print '*!* ATTEMPTING TO SAVE OUTPUT -- WAIT *!*'
wb.save('output - facebook.xlsx')

out_time = datetime.now().replace(microsecond=0)
duration = out_time-in_time
print '----------------------------> SENT'
print '       Completed in: ' + str(duration)
print 'w5748 <---------------------------'
time.sleep(2)
