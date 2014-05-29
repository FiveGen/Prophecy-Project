#INITIAL IMPORTS
from datetime import datetime
import time, random
import urllib
import mechanize
import re

#IMPORT EXCEL READ / WRITE / UTILIS LIBRARIES
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

#DEFINE TIME
pingtime = datetime.now().strftime('%Y-%m-%d %H:%M')
in_time = datetime.now().replace(microsecond=0)

#TITLE
print ' _    _  _____  ______ ___  _____ '
print '| |  | ||  ___||___  //   ||  _  |'
print '| |  | ||___ \    / // /| | \ V / '
print '| |/\| |    \ \  / // /_| | / _ \ '
print '\  /\  //\__/ /_/ / \___  || |_| |'
print ' \/  \/ \____/ \_/      |_/\_____/\n'

print '----------------------------> INFO'
print '* ' + pingtime
print '* C:\dignin\output - facebook.xls' + '\n'
print '----------------------------> LOG'


#OPEN OUTPUT.XLS
wb = load_workbook('output - facebook.xlsx')
ws = wb.get_sheet_by_name("sheet 1")

#BROWSER AGENT SETTINGS
browser = mechanize.Browser()
browser.set_handle_robots(False)
browser.addheaders = [('User-Agent', 'Chrome 3.2')]

#BROWSER DEBUG CODES - SET TRUE FOR ADDITIONAL INFO
browser.set_debug_http(False)
browser.set_debug_redirects(False)
browser.set_debug_responses(False)


#INITIAL LOGIN
browser.open("http://facebook.com/")
browser.select_form(nr=0)

#FACEBOOK CREDENTIALS (NEEDED TO BYPASS SPLASH)
browser.form['email'] = '????????????????'
browser.form['pass'] = '????????????????'
browser.submit()

#DEFINING URLS
LOL = 'http://www.facebook.com/leagueoflegends/'
HON = 'http://www.facebook.com/heroesofnewerth/'
DO2 = 'http://www.facebook.com/dota2/'

GW2 = 'http://www.facebook.com/GuildWars2'
XIV = 'http://www.facebook.com/FinalFantasyXIV'
WOW = 'http://www.facebook.com/Warcraft'

ESO = 'http://www.facebook.com/ElderScrollsOnline/'
EQN = 'https://www.facebook.com/everquestnext/'
WIS = 'https://www.facebook.com/WildStarOnline/'

#LIST FORM / POSITION
LIST = [LOL, HON, DO2, GW2, XIV, WOW, ESO, EQN, WIS]
N = 0

#EXCEL PARAMETERS
R = 1
C = 1
cell = ws.cell(row = R, column = C)

#REPEAT FOR LOOP
repeat = 'yes'
while N <= 8:

    #MOVE THE ROW
    try:
        while repeat == 'yes':
            cell = ws.cell(row = R, column = 1)
            if cell.data_type != 's':
                R += 1
            else:
                repeat = 'no'

    except IndexError:
        pass           

    #SLEEP TIMER
    zzz = random.randint(3,10)

    print '[' + str(N) + '] ' + LIST[N]
    print '  + REQUESTING IN ~' + str(zzz) + ' SECONDS'
    time.sleep(zzz)

    #BROWSER POINTING, HTML SEARCH
    browser.open(LIST[N])
    htmlback = browser.response().read()
    items=re.findall('</span><div class="fsm fwn fcg"><div class="fsm fwn fcg">.*likes ·',htmlback)
    for x in items:
        print '  - SCRAPE: ' + x

    #FILTER OUT NON-NUMERIC CHARACTERS        
    for char in " abcdefghijklmnopqrstuvwxyz,.'^?!/;:":  
        x = x.replace(char,'')
    for char in ' ABCDEFGHIJKLMNOPQRSTUVWXYZ=#$%")(][><ﾂｷ ·':  
        x = x.replace(char,'')

    print '  - REPORT: ' + x + '\n'
    i = int(x)
    N+=1

    #WRITE TO THE SPREADSHEET, MOVE THE COLUMN
    target_cell = ws.cell(row=R, column=C)
    target_cell.value = i
    C += 1

date_cell = ws.cell(row=R, column=0)
date_cell.value = pingtime

print '*!* ATTEMPTING TO SAVE OUTPUT -- WAIT *!*'
wb.save('output - facebook.xlsx')

out_time = datetime.now().replace(microsecond=0)
duration = out_time-in_time
print '----------------------------> SENT'
print '       Completed in: ' + str(duration)
print 'w5748 <---------------------------'
time.sleep(2)
