#INITIAL IMPORTS
from datetime import datetime
import time, random
import re
import csv
import os

#IMPORT EXCEL READ / WRITE / UTILIS LIBRARIES
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

#DEFINE TIME
pingtime = datetime.now().strftime('%Y-%m-%d %H:%M')
in_time = datetime.now().replace(microsecond=0)

#TITLE
print ' _    _  _____  ______ ___  _____ '
print '| |  | ||  ___||___  //   ||  _  |'
print '| |  | ||___ \    / // /| | \ V / '
print '| |/\| |    \ \  / // /_| | / _ \ '
print '\  /\  //\__/ /_/ / \___  || |_| |'
print ' \/  \/ \____/ \_/      |_/\_____/\n'

print '----------------------------> INFO'
print '* ' + pingtime
print '* \output - twitter.xlsx' + '\n\n'


#################################### TWEET DATA
print '----------------------------> TWEET DATA'

#OPEN OUTPUT.XLS
wb = load_workbook('output - twitter.xlsx')
ws = wb.get_sheet_by_name("sheet 1")

#TWEET LIST
LOL = 'twitout - LOL.csv'
HON = 'twitout - HON.csv'
DO2 = 'twitout - DO2.csv'

GW2 = 'twitout - GW2.csv'
XIV = 'twitout - XIV.csv'
WOW = 'twitout - WOW.csv'

ESO = 'twitout - ESO.csv'
EQN = 'twitout - EQN.csv'
WIS = 'twitout - WIS.csv'

#LIST FORM / POSITION
LIST = [LOL, HON, DO2, GW2, XIV, WOW, ESO, EQN, WIS]
N = 0

#EXCEL PARAMETERS
R = 1
C = 1
cell = ws.cell(row = R, column = C)

#REPEAT FOR LOOP
repeat = 'yes'
while N <= 8:

    #MOVE THE ROW UNTIL IT'S EMPTY
    try:
        while repeat == 'yes':
            cell = ws.cell(row = R, column = 1)
            if cell.data_type != 's':
                R += 1
            else:
                repeat = 'no'

    except IndexError:
        pass           

    #COUNTING OUTPUT
    row_count = sum(1 for row in csv.reader(open(LIST[N])))
    print ' ('+str(N)+') ACCESSING: ' + LIST[N]
    print ' > FOUND: ' + str(row_count) + ' TWEETS' + '\n'

    #WRITE TO THE SPREADSHEET, MOVE THE COLUMN
    target_cell = ws.cell(row=R, column=C)
    target_cell.value = row_count
    C += 1

    #IF THE FILE EXISTS (IT SHOULD) DELETE IT
    if os.path.isfile(LIST[N]):
        os.remove(LIST[N])
    else:
        pass

    #CREATE A BLANK SLATE WITH PINGTIME
    saveFile = open(LIST[N],'a')
    saveFile.write(pingtime+'\n')
    saveFile.close()

    #AFTER RECREATING THE FILE, MOVE ON
    N+=1

date_cell = ws.cell(row=R, column=0)
date_cell.value = pingtime

print '***********************************'
print '* SAVING TWEET DATA - PLEASE WAIT *'
wb.save('output - twitter.xlsx')
print '*           SUCCESS!              *'
print '***********************************' + '\n\n'


#################################### USERNAME DATA
print '----------------------------> USERNAME DATA'

#OPEN OUTPUT.XLS
wb = load_workbook('output - twitter.xlsx')
ws = wb.get_sheet_by_name("sheet 2")

#CHECKPOINT LISTS
LOLC = 'twitout - LOL - UN.txt'
HONC = 'twitout - HON - UN.txt'
DO2C = 'twitout - DO2 - UN.txt'

GW2C = 'twitout - GW2 - UN.txt'
XIVC = 'twitout - XIV - UN.txt'
WOWC = 'twitout - WOW - UN.txt'

ESOC = 'twitout - ESO - UN.txt'
EQNC = 'twitout - EQN - UN.txt'
WISC = 'twitout - WIS - UN.txt'


#LIST FORM / POSITION
LISTC = [LOLC, HONC, DO2C, GW2C, XIVC, WOWC, ESOC, EQNC, WISC]
N = 0
C = 1

#REPEAT FOR LOOP
repeat = 'yes'
while N <= 8:

    #COUNTING OUTPUT
    row_count = sum(1 for line in open(LISTC[N]))
    print ' ('+str(N)+') ACCESSING: ' + LISTC[N]
    print ' > FOUND ' + str(row_count) + ' UNIQUE USERNAMES' + '\n'

    #WRITE TO THE SPREADSHEET, MOVE THE COLUMN
    target_cell = ws.cell(row=R, column=C)
    target_cell.value = row_count
    C += 1

    #IF THE FILE EXISTS (IT SHOULD) DELETE IT
    if os.path.isfile(LISTC[N]):
        os.remove(LISTC[N])
    else:
        pass

    #RECREATE THE FILE WITH PINGTIME
    saveFile = open(LISTC[N],'a')
    saveFile.write(pingtime+'\n')
    saveFile.close()

    #NEXT LIST ITEM
    N+=1

date_cell = ws.cell(row=R, column=0)
date_cell.value = pingtime

print '***********************************'
print '*  SAVING U/N DATA - PLEASE WAIT  *'
wb.save('output - twitter.xlsx')
print '*           SUCCESS!              *'
print '***********************************' + '\n'



out_time = datetime.now().replace(microsecond=0)
duration = out_time-in_time
print '----------------------------> SENT'
print '       Completed in: ' + str(duration)
print 'w5748 <---------------------------'
time.sleep(2)
